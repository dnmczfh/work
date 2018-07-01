# coding: utf-8

#把目录下的文档名及相关信息存入电子表格，并打开文件

import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment,Font,colors
import win32ui
import configparser

def select_path(currentPath):
    '''
    通过使用对话窗口选择文件确定目录
    :return: 选择的路径名
    '''

    #设置起始路径,如果目标路径不存在，则选择当前路径
    if os.path.isdir(currentPath):
        os.chdir(currentPath)
    else:
        print('目录不存在:',currentPath)
        currentPath = os.getcwd()

    print("当前目录：%s"%(currentPath))

    dlg = win32ui.CreateFileDialog(1)  # 1表示打开文件对话框
    dlg.SetOFNInitialDir(currentPath)  # 设置打开文件对话框中的初始显示目录
    flag = dlg.DoModal()
    if flag == 2:
        print("未选择文件，程序退出")
        exit(1)

    filename = dlg.GetPathName()  # 获取选择的文件名称
    dlg.GetPathName()

    strSelectPath = os.path.dirname(filename)
    os.chdir(strSelectPath)

    currentPath = os.getcwd()
    print("工作目录：{}".format(currentPath))
    return currentPath

def list_dir_to_excel(currentPath,filename):
    '''
    把目录下的文档名及相关信息存入电子表格
    :param currentPath: 路径名
    :param filename: 存储的文件名
    :return: 0
    '''
    dirList =os.listdir(currentPath)
    fileList = []
    for eachFile in dirList:
        # 目录和文件使用不同的显示
        if os.path.isfile(eachFile):
            file = {'file':os.path.splitext(eachFile)[0],
                    'filename':eachFile,
                    'date': datetime.datetime.fromtimestamp(os.path.getctime(eachFile)).strftime("%Y-%m-%d"),
                    'link': currentPath+'\\'+eachFile}
            #print(file)
            fileList.append(file)

    print('共{}个文件'.format(len(fileList)))

    #按文件日期对列表排序
    fileList = sorted(fileList, key=lambda k: k.__getitem__('date'), reverse=False)

    # 实例一个Workbook
    wb = Workbook()

    # 字体
    ws = wb.active
    ws.title = '文件'
    ws.append(['存储目录：',currentPath])

    boldRedFont = Font(sz=16,name='Time New Roman',bold=True,color=colors.RED)
    ws['B1'].font = boldRedFont

    #写标题行且对中
    ws.append(['序号','文件名','文件日期','链接','地址'])
    strAlignment = Alignment(horizontal='center', vertical='center')
    for j in range(1,6):
        ws.cell(row=2, column=j).alignment = strAlignment

    id = 2
    for file in fileList:
        id = id +1

        ws['A'+ str(id)] = '=row()-2'
        ws['B'+ str(id)] = file['file']
        ws['C'+ str(id)] = file['date']
        ws['D' + str(id)] = '链接'
        ws['E'+ str(id)] = file['filename']
        link = file['link']
        ws['D' + str(id)].hyperlink = link

        # 设置数据垂直居中和水平居中
        ws['A'+ str(id)].alignment = strAlignment
        ws['D' + str(id)].alignment = strAlignment

    # 第1行行高
    ws.row_dimensions[1].height = 20
    # C列列宽
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 70
    ws.column_dimensions['C'].width = 11
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 70

    wb.save(filename)
    return 0

def readConfigFile(cfgFile):
    #使用全局变量
    global currentPath, Select, FileName
    #从配置文件中读取设置信息
    if os.path.isfile(cfgFile):#判断文件是否存在
        #从文件中读取配置信息
        try:
            config.read(cfgFile)
            currentPath = config['Path']['Target']
            Select = config['Path']['Select']
            FileName = config['File']['Name']
        except Exception as e:
            print('错误信息:', e)
            writeConfigFile(cfgFile)
    else:
        #生成新的配置文件
        writeConfigFile(cfgFile)
    return(0)

def writeConfigFile(cfgFile):
    print(cfgFile, '不存在')
    config.add_section('Path')
    config['Path']['Target'] = os.getcwd()
    config['Path']['Select'] = 'Y'
    config.add_section('File')
    config['File']['Name'] = 'file_list.xlsx'
    config.write(open(cfgFile, "w"))
    print(cfgFile, '已经重新生成')
    return(0)


if __name__ == '__main__':
    #默认设置
    currentPath = os.getcwd()
    Select = 'Y'
    FileName = 'file_list.xlsx'

    # 记录配置文件的全路径名
    cfgFile = os.path.abspath('path.ini')
    config = configparser.ConfigParser()

    readConfigFile(cfgFile)

    # 设置需要处理的目录
    if os.path.isdir(currentPath):
        os.chdir(currentPath)
    else:
        # 使用对话框选择路径
        currentPath = select_path(currentPath)
        # 不再需要重新选择路径
        Select = 0

    if Select == 1 or Select == 'Y' or Select == 'y':
        # 使用对话框选择路径
        currentPath = select_path(currentPath)
        #把新选择写入配置文件
        config['Path']['Target'] = currentPath
        config.write(open(cfgFile, "w"))

    #把目录下的文档名及相关信息存入电子表格
    list_dir_to_excel(currentPath, FileName)

    #打开文件
    os.startfile(FileName)
