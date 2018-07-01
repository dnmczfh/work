#python 3.6
#读写excel文件

import openpyxl


# #创建文件
# from openpyxl import Workbook
# # 实例一个Workbook
# wb = Workbook()
# # 激活工作表
# ws = wb.active
# ws.title='Sheet1'
# # 数据可以直接赋值给单元格
# ws['A1'] = 42
# # 在下一行追加一行1,2,3
# ws.append([1, 2, 3])
# ws.append([3, 2, 1])
# ws.append([4, 5, 6])
# ws.append([7, 8, 9])
# # Python类型可以自动转换成Excel类型
# import datetime
# ws['A3'] = datetime.datetime.now()
# # 保存文件
# wb.save("sample.xlsx")

import datetime
print(datetime.datetime.now())
wb = openpyxl.load_workbook("sample.xlsx")
mySheet = wb.create_sheet('mySheet')
print(wb.sheetnames)

for sheet in wb:
    print(sheet.title)

print(datetime.datetime.now())
for sheetname in wb.sheetnames:
    print(sheetname)

ws = wb.active
ws = wb.get_sheet_by_name('Sheet2')
#取单元格的值
# c = ws['A3']
# print('行{}列{}的 值 是{}:'.format(c.row,c.column,c.value))
# print('单元格{}的 值 是{}:'.format(c.coordinate,c.value))
# print(ws.cell(row=3,column=1).value)

#按行或列从表格中取数
#按列取
colC = ws['C']
print(colC)
#按行取
row3 = ws[3]
print(row3)

#取行、列的范围取
col_range = ws['B:C']
row_range = ws[2:4]

for col in col_range :
    for cell in col:
        print(cell.value)

for row in row_range:
    for cell in row:
        print(cell.value)

#取特定范围的区间
iter_range = ws.iter_cols(2,4,3,5)
for col in iter_range:
    for cell in col:
        print(cell)
cell_range = ws['A1:C3']
for rowOfCellObject in cell_range:
    for cellObject in rowOfCellObject:
        print(cellObject.coordinate, cellObject.value)
    print('------------End of Row-------------')

print('{}行*{}列'.format(ws.max_row,ws.max_column))
